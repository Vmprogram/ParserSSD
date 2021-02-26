import json
import requests
from bs4 import BeautifulSoup as bs
import lxml
import os
import sys
from time import time, sleep
from fake_useragent import UserAgent
import re
import openpyxl
import time
from Parser import *
'''
Нужно получить:
	Код
	Имя раздела
	Наименование
	Ед.изм.
	Цена розн. без НДС
	Цена опт. без НДС
	Размер опт.
	Дистриб. скидки
	Кол-во в упак.
	Производство (в среднем)
	Дата утв. цены

	with open(f'{Urls.index(url_main_page)}.html', 'w', encoding='utf-8') as output_file:
			output_file.write(page.text)

'''
Urls = ['https://www.ssd.ru/mufty-dlya-opticheskikh-kabeley','https://www.ssd.ru/stantsionnaya-storona-vols','https://www.ssd.ru/abonentskaya-storona-vols','https://www.ssd.ru/pribory-i-instrumenty-dlya-vols','https://www.ssd.ru/armatura-dlya-podveski-opticheskikh-kabeley']
Agent = UserAgent().random
s = requests.Session()
headers = {'User-Agent': Agent}
Items = {}
#Items = {'MainClass': {'SubClass': [{'Code': '0-0-0-0', 'Name': 'Test', 'unitValue': 'ШТ','PriceCountDNDS': 100,'PriceOptDNDS': 95,'CountOpt': 10,'Discount': 10,'Count': 10,'CreateAVG': 10,'datePrice': '28.01.2020'},{'Code': '0-0-0-0', 'Name': 'Test', 'CountType': 'ШТ','PriceCountDNDS': 100,'PriceOptDNDS': 95,'CountOpt': 10,'Discount': 10,'Count': 10,'CreateAVG': 10,'datePrice': '28.01.2020'}]}}
#soup = bs(page.text, 'lxml')

def getMainPage(url):
	print(f'Открываю: {url}')
	page = s.get(url,headers=headers)
	return(page)

def Main():
	global Items

	for url_main_page in Urls:
		page = getMainPage(url_main_page)
		soup = bs(page.text, 'lxml')
		NameMainClass = getInfoMainPage(soup) #Items[NameMainClass]
		Items[NameMainClass] = {}
		SubClassNames = getAllSubClass(soup)
		for SubClassName in SubClassNames:
			HrefItems = getItemsList(SubClassNames[SubClassName])
			for HrefItem in HrefItems:
				#infoItem = getInfoItems(HrefItem)
				infoItem = Parser.getItems(HrefItem,s,headers)
				print(f'Работаю во вкладке: {NameMainClass}->{SubClassName}')
				if infoItem == False:
					continue
				elif infoItem != None:
					if not Items[NameMainClass].get(SubClassName):
						Items[NameMainClass][SubClassName] = []
					Items[NameMainClass][SubClassName].append(infoItem)
					
				
	Excel.start(Items)
				# if i//10 == 0:
				# 	print('Пауза 10 секунд')
				# 	time.sleep(10)

class Excel:
	def start(Items):
		excel,sheet = Excel.CreateHead()
		excel,sheet = Excel.fulling(excel,sheet,Items)
		excel.save('mybook.xlsx')
		excel.close()

	def CreateHead():
		excel = openpyxl.Workbook()
		sheet = excel.active

		sheet.merge_cells('A1:C3')
		sheet['A1'] = 'ЗАО "Связьстройдеталь"'

		sheet.merge_cells('D1:E1')
		sheet['D1'] = 'http://www.ssd.ru'

		sheet.merge_cells('D2:E2')
		sheet['D2'] = 'e-mail: mail@ssd.ru'

		sheet.merge_cells('F1:J1')
		sheet['F1'] = '115088, Москва, ул. Южнопортовая, 7а'

		sheet.merge_cells('F2:J2')
		sheet['F2'] = '(495) 786-34-34 - телефон'

		sheet.merge_cells('F3:J3')
		sheet['F3'] = '(495) 786-34-32 - факс'

		sheet.merge_cells('A4:J4')
		sheet['A4'] = '3. Монтаж оптических кабелей связи'

		
		return(excel,sheet)

	def fulling(excel,sheet,Items):
		Position = 5
		Dict = 3.1
		for NameMainClass in Items:
			sheet.merge_cells(f'A{Position}:J{Position}')
			sheet[f'A{Position}'] = f'{Dict}{NameMainClass}'
			Position += 1
			for NameSubClass in Items[NameMainClass]:
				sheet.merge_cells(f'A{Position}:J{Position}')
				sheet[f'A{Position}'] = f'{NameSubClass}'
				Position += 1
				#-------------------------------
				sheet[f'A{Position}'] = 'Код'
				sheet[f'B{Position}'] = 'Имя'
				sheet[f'C{Position}'] = 'Ед. Изм'
				sheet[f'D{Position}'] = 'Цены'
				sheet[f'E{Position}'] = 'Размер опт.'
				sheet[f'F{Position}'] = 'Оптовая скидка'
				sheet[f'G{Position}'] = 'Дистриб. скидки'
				sheet[f'H{Position}'] = 'Склад'
				sheet[f'I{Position}'] = 'Время отправки'
				sheet[f'J{Position}'] = 'Характеристики'
				sheet[f'K{Position}'] = 'Комплектация'
				sheet[f'L{Position}'] = 'Логистические параметры'
				sheet[f'M{Position}'] = 'Маркеры'
				sheet[f'N{Position}'] = 'Инструкции'
				sheet[f'O{Position}'] = 'Описание'
				sheet[f'P{Position}'] = 'Декларации'
				sheet[f'Q{Position}'] = 'Фото'
				Position += 1
				#------------------------------
				for Item in Items[NameMainClass][NameSubClass]:
					code = Item[0]
					name = Item[1]
					if Item[2] == None:
						edIzm = 'Нет информации'
					else:
						edIzm = Item[2]

					if Item[3] == None:
						prices = 'Нет информации'
					else:
						prices = ''
						for a in Item[3]:
							prices += f'{a}\n' 

					if Item[4] == None:
						sizeOpt = 'Нет информации'
					else:
						sizeOpt = Item[4]

					if Item[5] == None:
						optDiscount = 'Нет информации'
					else:
						optDiscount = Item[5]

					if Item[6] == None:
						distripDscount = 'Нет информации'
					else:
						distripDscount = Item[6]

					if Item[7] == None:
						sklad = 'Нет информации'
					else:
						sklad = ''
						for a in Item[7]:
							sklad += f'{a}\n' 

					if Item[8] == None:
						timeSend = 'Нет информации'
					else:
						timeSend = Item[8]

					if Item[9] == None:
						Charectes = 'Нет информации'
					else:
						Charectes = ''
						for a in Item[9]:
							Charectes += f'{a}\n' 

					if Item[10] == None:
						Complects = 'Нет информации'
					else:
						Complects = ''
						for a in Item[10]:
							Complects += f'{a}\n' 

					if Item[11] == None:
						Logistick = 'Нет информации'
					else:
						Logistick = ''
						for a in Item[11]:
							Logistick += f'{a}\n' 

					if Item[12] == None:
						Marker = 'Нет информации'
					else:
						Marker = ''
						for a in Item[12]:
							Marker += f'{a}\n' 

					if Item[13] == None:
						Instruct = 'Нет информации'
					else:
						Instruct = ''
						for a in Item[13]:
							Instruct += f'{a}\n' 

					if Item[14] == None:
						description = 'Нет информации'
					else:
						description = Item[14]

					if Item[15] == None:
						Declaracion = 'Нет информации'
					else:
						Declaracion = ''
						for a in Item[15]:
							Declaracion += f'{a}\n' 

					if Item[16] == None:
						Photos = 'Нет информации'
					else:
						Photos = ''
						for a in Item[16]:
							Photos += f'{a}\n' 

					'''
					code 'Код'
					name 'Имя'
					edIzm 'Ед. Изм'
					prices 'Цены'
					sizeOpt 'Размер опт.'
					optDiscount 'Оптовая скидка'
					distripDscount 'Дистриб. скидки'
					sklad 'Склад'
					timeSend 'Время отправки'
					Charectes 'Характеристики'
					Complects 'Комплектация'
					Logistick 'Логистические параметры'
					Marker 'Маркеровка'
					Instruct 'Инструкция'
					description 'Описание'
					Declaracion 'Декларации'
					Photos 'Фото'
					
					'''
					#-------------------------------
					sheet[f'A{Position}'] = code
					sheet[f'B{Position}'] = name
					sheet[f'C{Position}'] = edIzm
					sheet[f'D{Position}'] = prices
					sheet[f'E{Position}'] = sizeOpt
					sheet[f'F{Position}'] = optDiscount
					sheet[f'G{Position}'] = distripDscount
					sheet[f'H{Position}'] = sklad
					sheet[f'I{Position}'] = timeSend
					sheet[f'J{Position}'] = Charectes
					sheet[f'K{Position}'] = Complects
					sheet[f'L{Position}'] = Logistick
					sheet[f'M{Position}'] = Marker
					sheet[f'N{Position}'] = Instruct
					sheet[f'O{Position}'] = description
					sheet[f'P{Position}'] = Declaracion
					sheet[f'Q{Position}'] = Photos
					Position += 1
					#------------------------------

					

		Dict += 0.1
		return(excel,sheet)

def Test():
	url = '/kronshteyn-dlya-podveski-mufty-mpo-sh1-standart-ssd'
	getInfoItems(url)

# def getInfoItems(url):
# 	print(f'Открываю: https://www.ssd.ru{url}')
# 	page = s.get(f'https://www.ssd.ru{url}',headers=headers)
# 	soup = bs(page.text, 'lxml')
# 	status = soup.find('div',{'class':'widget__title'}).text
# 	name = soup.find('h1',{'class':'page-title'}).get('content') #Название товара
# 	#--------------------------------------------------------

# 	code = soup.find('span',{'class':'copy-to-buffer'})
# 	code = code.find('strong').text #Код
# 	#--------------------------------------------------------

	
# 	if status.find('Товар под заказ') == -1:
# 		PriceList = []
# 		blockPrices = soup.findAll('td',{'class':'price-cell'})
# 		for blockPrice in blockPrices:
# 			PriceList.append(blockPrice.find('span').text)
# 		#--------------------------------------------------------

# 		unitValue = soup.find('div',{'class':'unit-value'}).text #Ед. Изм
# 		#--------------------------------------------------------

# 		CountOpt = soup.findAll('tr',{'class':'way'})[0]
# 		CountOpt = CountOpt.findAll('td')[1].text
# 		CountOpt = re.sub('[^0-9]', '', CountOpt)#Размер опт.
# 		#--------------------------------------------------------

# 		countBlock = len(soup.findAll('table',{'class':'price-table-content'}))
# 		if countBlock == 8:
# 			Discount = soup.findAll('table',{'class':'price-table-content'})[2]
# 			Discount = Discount.find('span').text #Дистриб. скидки
# 		elif countBlock == 10:
# 			Discount = soup.findAll('table',{'class':'price-table-content'})[3]
# 			Discount = Discount.find('span').text #Дистриб. скидки
# 		else:
# 			print('Не нашел информацию о скидки')
# 			Discount = 'Нет информации'
# 		#--------------------------------------------------------

# 		TimeSending = soup.find('div',{'class':'avr-post-block'}).text
# 		TimeSending = re.sub(" +", " ", TimeSending).replace('\n','')
# 		TimeSending = TimeSending[1:] #Информация о времени отправки
# 		#--------------------------------------------------------
# 		CountSklad = soup.findAll('div',{'class':'accord__body'})[0]
# 		CountSkladTbody = CountSklad.find('tbody')
# 		CountSkladTRs = CountSklad.findAll('tr')
# 		SkladInfo = {}
# 		for CountSkladTR in CountSkladTRs:
# 			nameSkald = CountSkladTR.findAll('td')[0].text
# 			countSkald = CountSkladTR.findAll('td')[1].text
# 			SkladInfo[nameSkald] = countSkald #Информация о наличие на складе

# 		#--------------------------------------------------------
# 		BlocksTrDiscount = soup.findAll('tr')
# 		for BlockTrDiscount in BlocksTrDiscount:
# 			if len(BlockTrDiscount.findAll('td')) > 1 and len(BlockTrDiscount.findAll('td')) < 5:
# 				if BlockTrDiscount.findAll('td')[0].text.find('Оптовая скидка') != -1:
# 					percentDiscount = BlockTrDiscount.findAll('td')[1].find('span').text
# 					percentDiscount = re.sub(" +", " ", percentDiscount).replace('\n','') #Оптовая скидка
# 					break
# 		#--------------------------------------------------------
# 		if len(soup.findAll('div',{'class':'accord__body'})) == 9:
# 			CharectesBlock = soup.findAll('div',{'class':'accord__body'})[2]
# 			CharectesBlock = CharectesBlock.find('table')
# 			CharectesBlockTRs = CharectesBlock.findAll('tr')
# 			CharectesInfo = {}
# 			for CharectesBlockTR in CharectesBlockTRs:
# 				nameCharacters = CharectesBlockTR.findAll('td')[0].text
# 				AnswerCharacters = CharectesBlockTR.findAll('td')[1].text
# 				CharectesInfo[nameCharacters] = AnswerCharacters #Характеристики товара

# 		elif len(soup.findAll('div',{'class':'accord__body'})) == 7:
# 			CharectesBlock = soup.findAll('div',{'class':'accord__body'})[2]
# 			CharectesBlock = CharectesBlock.find('table')
# 			CharectesBlockTRs = CharectesBlock.findAll('tr')
# 			CharectesInfo = {}
# 			for CharectesBlockTR in CharectesBlockTRs:
# 				nameCharacters = CharectesBlockTR.findAll('td')[0].text
# 				AnswerCharacters = CharectesBlockTR.findAll('td')[1].text
# 				CharectesInfo[nameCharacters] = AnswerCharacters #Характеристики товара
# 		else:
# 			CharectesInfo = None
# 		#--------------------------------------------------------

# 		if len(soup.findAll('div',{'class':'accord__body'})) == 9:
# 			ComplectBlock = soup.findAll('div',{'class':'accord__body'})[3]
# 			ComplectBlock = ComplectBlock.find('table')
# 			ComplectBlockTRs = ComplectBlock.findAll('tr')
# 			ComplectInfo = {}
# 			for ComplectBlockTR in ComplectBlockTRs:
# 				nameComplect = ComplectBlockTR.findAll('td')[0].text
# 				AnswerComplect = ComplectBlockTR.findAll('td')[1].text
# 				ComplectInfo[nameComplect] = AnswerComplect #Комплектация товара

# 		elif len(soup.findAll('div',{'class':'accord__body'})) == 7:
# 			ComplectBlock = soup.findAll('div',{'class':'accord__body'})[3]
# 			ComplectBlock = ComplectBlock.find('table')
# 			ComplectBlockTRs = ComplectBlock.findAll('tr')
# 			ComplectInfo = {}
# 			for ComplectBlockTR in ComplectBlockTRs:
# 				nameComplect = ComplectBlockTR.findAll('td')[0].text
# 				AnswerComplect = ComplectBlockTR.findAll('td')[1].text
# 				ComplectInfo[nameComplect] = AnswerComplect #Комплектация товара

# 		elif len(soup.findAll('div',{'class':'accord__body'})) == 5:
# 			ComplectBlock = soup.findAll('div',{'class':'accord__body'})[2]
# 			ComplectBlock = ComplectBlock.find('table')
# 			ComplectBlockTRs = ComplectBlock.findAll('tr')
# 			ComplectInfo = {}
# 			for ComplectBlockTR in ComplectBlockTRs:
# 				nameComplect = ComplectBlockTR.findAll('td')[0].text
# 				AnswerComplect = ComplectBlockTR.findAll('td')[1].text
# 				ComplectInfo[nameComplect] = AnswerComplect #Комплектация товара

# 		else:
# 			ComplectInfo = None

# 		#--------------------------------------------------------
# 		if len(soup.findAll('div',{'class':'accord__body'})) == 9:
# 			LogistickBlock = soup.findAll('div',{'class':'accord__body'})[4]
# 			LogistickBlock = LogistickBlock.find('table')
# 			LogistickBlockTRs = LogistickBlock.findAll('tr')
# 			LogistickInfo = {}
# 			for LogistickBlockTR in LogistickBlockTRs:
# 				nameLogistick = LogistickBlockTR.findAll('td')[0].text
# 				AnswerLogistick = LogistickBlockTR.findAll('td')[1].text
# 				LogistickInfo[nameLogistick] = AnswerLogistick #Логистические параметры товара

# 		elif len(soup.findAll('div',{'class':'accord__body'})) == 7:
# 			LogistickBlock = soup.findAll('div',{'class':'accord__body'})[4]
# 			LogistickBlock = LogistickBlock.find('table')
# 			LogistickBlockTRs = LogistickBlock.findAll('tr')
# 			LogistickInfo = {}
# 			for LogistickBlockTR in LogistickBlockTRs:
# 				nameLogistick = LogistickBlockTR.findAll('td')[0].text
# 				AnswerLogistick = LogistickBlockTR.findAll('td')[1].text
# 				LogistickInfo[nameLogistick] = AnswerLogistick

# 		elif len(soup.findAll('div',{'class':'accord__body'})) == 5:
# 			LogistickBlock = soup.findAll('div',{'class':'accord__body'})[3]
# 			LogistickBlock = LogistickBlock.find('table')
# 			LogistickBlockTRs = LogistickBlock.findAll('tr')
# 			LogistickInfo = {}
# 			for LogistickBlockTR in LogistickBlockTRs:
# 				nameLogistick = LogistickBlockTR.findAll('td')[0].text
# 				AnswerLogistick = LogistickBlockTR.findAll('td')[1].text
# 				LogistickInfo[nameLogistick] = AnswerLogistick
# 		else:
# 			LogistickInfo = None
# 		#--------------------------------------------------------
# 		if len(soup.findAll('div',{'class':'article'})) > 1:
# 			ArticleBlock = soup.findAll('div',{'class':'article'})[0]
# 			ArticleBlock = soup.find('div',{'itemprop':'description'}).text
# 			Article = ArticleBlock.replace('Перейти к сопутствующим товарам', '')	
# 			Article = re.sub(" +", " ", Article).replace('\n','') #Описание товара
# 		else:
# 			Article = None
# 		#--------------------------------------------------------
# 		if len(soup.findAll('img',{'class':'swiper-lazy'})) > 0:
# 			MainPhoto = soup.find('img',{'class':'swiper-lazy'}).get('src')
# 			Photos = []	
# 			PhotoBlock = soup.findAll('div',{'class':'swiper-container'})[1]
# 			PhotoBlock = PhotoBlock.findAll('img',{'class':'swiper-lazy'})
# 			for SubPhoto in PhotoBlock:
# 				Photos.append(SubPhoto.get('src'))
# 		else:
# 			Photos = None
# 			MainPhoto = None
# 		#---------------------------------------------------------
# 		print(len(soup.findAll('div',{'class':'accord__body'})))
# 		item = [code,name,unitValue,PriceList,CountOpt,percentDiscount,Discount,SkladInfo,TimeSending,CharectesInfo,ComplectInfo,LogistickInfo,Article,MainPhoto,Photos]
# 		#print(f'Имя: {name}\nКод: {code}\nЕд. Изм: {unitValue}\nЦены: {PriceList}\nРазмер опт.: {CountOpt}\nОптовая скидка: {percentDiscount}\nДистриб. скидки: {Discount}\nСклад: {SkladInfo}\nВремя отправки: {TimeSending}\nХарактеристика: {CharectesInfo}\nКомплект:{ComplectInfo}\nЛогистик: {LogistickInfo}\nОписание: {Article}')
# 		return(item)
# 	else:
# 		PriceList = None
# 		CountOpt = None
# 		Discount = None
# 		TimeSending = None
# 		SkladInfo = None
# 		percentDiscount = None
# 		unitValue = None
# 		print(len(soup.findAll('div',{'class':'accord__body'})))
# 		if len(soup.findAll('div',{'class':'accord__body'})) == 5:
# 			CharectesBlock = soup.findAll('div',{'class':'accord__body'})[0]
# 			CharectesBlock = CharectesBlock.find('table')
# 			CharectesBlockTRs = CharectesBlock.findAll('tr')
# 			CharectesInfo = {}
# 			for CharectesBlockTR in CharectesBlockTRs:
# 				nameCharacters = CharectesBlockTR.findAll('td')[0].text
# 				AnswerCharacters = CharectesBlockTR.findAll('td')[1].text
# 				CharectesInfo[nameCharacters] = AnswerCharacters #Характеристики товара
# 		else:
# 			print('ДРУГОЕ КОЛЛ-ВО ЭЛЕМЕНТОВ В СПИСКЕ!!!!')
# 			CharectesInfo = None

# 		if len(soup.findAll('div',{'class':'accord__body'})) == 5:
# 			ComplectBlock = soup.findAll('div',{'class':'accord__body'})[1]
# 			ComplectBlock = ComplectBlock.find('table')
# 			ComplectBlockTRs = ComplectBlock.findAll('tr')
# 			ComplectInfo = {}
# 			for ComplectBlockTR in ComplectBlockTRs:
# 				nameComplect = ComplectBlockTR.findAll('td')[0].text
# 				AnswerComplect = ComplectBlockTR.findAll('td')[1].text
# 				ComplectInfo[nameComplect] = AnswerComplect #Комплектация товара
# 		else:
# 			print('ДРУГОЕ КОЛЛ-ВО ЭЛЕМЕНТОВ В СПИСКЕ!!!!')
# 			ComplectInfo = None

# 		if len(soup.findAll('div',{'class':'accord__body'})) == 5:
# 			LogistickBlock = soup.findAll('div',{'class':'accord__body'})[2]
# 			LogistickBlock = LogistickBlock.find('table')
# 			LogistickBlockTRs = LogistickBlock.findAll('tr')
# 			LogistickInfo = {}
# 			for LogistickBlockTR in LogistickBlockTRs:
# 				nameLogistick = LogistickBlockTR.findAll('td')[0].text
# 				AnswerLogistick = LogistickBlockTR.findAll('td')[1].text
# 				LogistickInfo[nameLogistick] = AnswerLogistick #Логистические параметры товара
# 		else:
# 			print('ДРУГОЕ КОЛЛ-ВО ЭЛЕМЕНТОВ В СПИСКЕ!!!!')		
# 			LogistickInfo = None

# 		ArticleBlock = soup.findAll('div',{'class':'article'})[0]
# 		ArticleBlock = soup.find('div',{'itemprop':'description'}).text
# 		Article = ArticleBlock.replace('Перейти к сопутствующим товарам', '')	
# 		Article = re.sub(" +", " ", Article).replace('\n','') #Описание товара
# 		#--------------------------------------------------------
# 		MainPhoto = soup.find('img',{'class':'swiper-lazy'}).get('src')
# 		Photos = []
# 		PhotoBlock = soup.findAll('div',{'class':'swiper-container'})[1]
# 		PhotoBlock = PhotoBlock.findAll('img',{'class':'swiper-lazy'})
# 		for SubPhoto in PhotoBlock:
# 			Photos.append(SubPhoto.get('src'))
# 		#---------------------------------------------------------
# 		item = [code,name,unitValue,PriceList,CountOpt,percentDiscount,Discount,SkladInfo,TimeSending,CharectesInfo,ComplectInfo,LogistickInfo,Article,MainPhoto,Photos]
# 		return(item)

def getItemsList(url):
	print(f'Открываю: https://www.ssd.ru{url}')
	page = s.get(f'https://www.ssd.ru{url}',headers=headers)
	soup = bs(page.text, 'lxml')
	blockHrefItems = []
	FullBlock = soup.find('div',{'class':'items-wrap'})
	FullBlock = FullBlock.findAll('div',{'class':'item'})
	for Block in FullBlock:
		Href = Block.find('a').get('href')
		blockHrefItems.append(Href)
	return(blockHrefItems)

def getAllSubClass(soup):
	blockHref = {}
	FullBlock = soup.find('div',{'class':'cats-wrap'})
	FullBlock = FullBlock.findAll('div',{'class':'cat'})
	for Block in FullBlock:
		Href = Block.find('a',{'class':'cat__in'}).get('href')
		name = Block.find('img').get('alt')
		blockHref[name] = Href
	return(blockHref)

def getInfoMainPage(soup):
	#Получение имени
	nameClass = soup.find('div',{'class':'page-head'}).find('h1').text
	nameClass = re.sub(" +", " ", nameClass).replace('\n','')
	nameClass = nameClass[1:]
	#-----------------
	return(nameClass)


Main()